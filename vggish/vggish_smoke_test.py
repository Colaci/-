# -*- coding: utf-8 -*-
# Copyright 2017 The TensorFlow Authors All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================

"""A smoke test for VGGish.
This is a simple smoke test of a local install of VGGish and its associated
downloaded files. We create a synthetic sound, extract log mel spectrogram
features, run them through VGGish, post-process the embedding ouputs, and
check some simple statistics of the results, allowing for variations that
might occur due to platform/version differences in the libraries we use.
Usage:
- Download the VGGish checkpoint and PCA parameters into the same directory as
  the VGGish source code. If you keep them elsewhere, update the checkpoint_path
  and pca_params_path variables below.
- Run:
  $ python vggish_smoke_test.py
"""

from __future__ import print_function

import numpy as np
import tensorflow.compat.v1 as tf
tf.disable_v2_behavior()

import vggish_input
import vggish_params
import vggish_postprocess
import vggish_slim
import os
from openpyxl import load_workbook
 

print('\nTesting your install of VGGish\n')

# Paths to downloaded VGGish files.
checkpoint_path = 'vggish_model.ckpt'
pca_params_path = 'vggish_pca_params.npz'

# Relative tolerance of errors in mean and standard deviation of embeddings.
rel_error = 0.1  # Up to 10%

# Generate a 1 kHz sine wave at 44.1 kHz (we use a high sampling rate
# to test resampling to 16 kHz during feature extraction).

# Produce a batch of log mel spectrogram examples.
wb=load_workbook(r'data_music.xlsx')
sheet=wb.active
path='161-166'
files=os.listdir(path)
for i in range (162,167):
  for x in files:
    if(str(i)+'.' in x):
      input_batch = vggish_input.wavfile_to_examples('161-166\\'+x)
	 
      #print('Log Mel Spectrogram example: ', input_batch[0])

# Define VGGish, load the checkpoint, and run the batch through the model to
# produce embeddings.
  with tf.Graph().as_default(), tf.Session() as sess:
    vggish_slim.define_vggish_slim()
    vggish_slim.load_vggish_slim_checkpoint(sess, checkpoint_path)

    features_tensor = sess.graph.get_tensor_by_name(
      vggish_params.INPUT_TENSOR_NAME)
    embedding_tensor = sess.graph.get_tensor_by_name(
      vggish_params.OUTPUT_TENSOR_NAME)
    [embedding_batch] = sess.run([embedding_tensor],
                               feed_dict={features_tensor: input_batch})
    print('VGGish embedding: ', embedding_batch[0])
 

# Postprocess the results to produce whitened quantized embeddings.
  pproc = vggish_postprocess.Postprocessor(pca_params_path)
  postprocessed_batch = pproc.postprocess(embedding_batch)
  print('Postprocessed VGGish embedding: ', postprocessed_batch[0])
  for n in range(8,136):
    sheet.cell(row=i+1,column=n).value=postprocessed_batch[0][n-8]

  wb.save(r'data_music.xlsx')
print('\nLooks Good To Me!\n')